#!/usr/bin/env python3
"""Format and mechanically check a pin assignment workbook.

This script intentionally does not infer pin assignments. It styles workbook
sheets and creates/refreshes a mechanical checks sheet from existing columns,
including duplicate nets/pins, blank pin/net mismatches, and source ID gaps.
"""

from __future__ import annotations

import argparse
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DEFAULT_SKIP_SHEETS = {
    "Inputs",
    "Sources",
    "Checks",
    "Review_Notes",
    "Change_Log",
    "Mechanical_Checks",
}
EXACT_NET_HEADERS = {
    "net",
    "net name",
    "net_name",
    "schematic net",
    "schematic net name",
    "lpddr5 netname",
    "lpddr5 net name",
}
EXCLUDE_NET_SUBSTRINGS = ("class", "group", "type", "role", "category")
SOURCE_HEADERS = {"source", "source id", "source ids", "sources", "source_id", "source_ids"}
SOURCE_REVIEW_MARKERS = {"tbd-source", "conflict"}
INTENTIONAL_DUPLICATE_HEADERS = {
    "intentional duplicate marker",
    "intentional duplicate",
    "intentional_duplicate",
    "duplicate ok",
    "duplicate_ok",
}
FPGA_PIN_SUBSTRINGS = ("fpga pin", "soc pin", "device pin")
PERIPHERAL_PIN_SUBSTRINGS = (
    "peripheral pin",
    "peripheral ball",
    "memory pin",
    "memory ball",
    "mapped peripheral pin",
    "mapped peripheral pin/ball",
)
GENERIC_PIN_EXACT_HEADERS = {"pin", "pin number", "pin/ball", "ball"}
EXCLUDE_PIN_SUBSTRINGS = ("pin name", "signal name", "vendor signal")
TRUE_MARKERS = {"y", "yes", "true", "1", "ok", "intentional"}


def parse_cols(value: str | None) -> set[int]:
    if not value:
        return set()
    cols: set[int] = set()
    for item in value.split(","):
        item = item.strip()
        if not item:
            continue
        if item.isdigit():
            cols.add(int(item))
        else:
            col = 0
            for ch in item.upper():
                if not ("A" <= ch <= "Z"):
                    raise ValueError(f"Invalid column: {item}")
                col = col * 26 + ord(ch) - ord("A") + 1
            cols.add(col)
    return cols


def normalize_header(value: object) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def infer_exact_columns(headers: Iterable[str], allowed: set[str]) -> set[int]:
    cols = set()
    for idx, header in enumerate(headers, start=1):
        h = normalize_header(header)
        if h in allowed:
            cols.add(idx)
    return cols


def infer_net_columns(headers: Iterable[str]) -> set[int]:
    cols = set()
    for idx, header in enumerate(headers, start=1):
        h = normalize_header(header)
        if any(blocked in h for blocked in EXCLUDE_NET_SUBSTRINGS):
            continue
        if h in {normalize_header(item) for item in EXACT_NET_HEADERS}:
            cols.add(idx)
    return cols


def infer_pin_columns(headers: Iterable[str], substrings: tuple[str, ...]) -> set[int]:
    cols = set()
    for idx, header in enumerate(headers, start=1):
        h = normalize_header(header)
        if any(blocked in h for blocked in EXCLUDE_PIN_SUBSTRINGS):
            continue
        if any(item in h for item in substrings):
            cols.add(idx)
    return cols


def infer_generic_pin_columns(headers: Iterable[str]) -> set[int]:
    cols = set()
    for idx, header in enumerate(headers, start=1):
        h = normalize_header(header)
        if h in {normalize_header(item) for item in GENERIC_PIN_EXACT_HEADERS}:
            cols.add(idx)
    return cols


def nonempty(value: object) -> bool:
    return value is not None and str(value).strip() != ""


def truthy_marker(value: object) -> bool:
    return str(value or "").strip().lower() in TRUE_MARKERS


def style_sheet(ws) -> None:
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
    for row in ws.iter_rows():
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col_idx in range(1, ws.max_column + 1):
        width = 12
        for row_idx in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is not None:
                width = max(width, min(42, len(str(value)) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


def collect_checks(wb, net_cols: set[int], pin_cols: set[int], skip_sheets: set[str]) -> list[list[object]]:
    checks: list[list[object]] = []
    for ws in wb.worksheets:
        if ws.title in skip_sheets:
            continue
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        sheet_net_cols = net_cols or infer_net_columns(headers)
        sheet_pin_cols = pin_cols or infer_generic_pin_columns(headers)
        fpga_pin_cols = pin_cols or infer_pin_columns(headers, FPGA_PIN_SUBSTRINGS)
        peripheral_pin_cols = pin_cols or infer_pin_columns(headers, PERIPHERAL_PIN_SUBSTRINGS)
        all_pin_cols = set().union(sheet_pin_cols, fpga_pin_cols, peripheral_pin_cols)
        source_cols = infer_exact_columns(headers, {normalize_header(item) for item in SOURCE_HEADERS})
        duplicate_marker_cols = infer_exact_columns(headers, {normalize_header(item) for item in INTENTIONAL_DUPLICATE_HEADERS})

        nets: list[tuple[int, int, str]] = []
        pins: list[tuple[int, int, str]] = []
        duplicate_rows = {
            row_idx
            for row_idx in range(2, ws.max_row + 1)
            if any(truthy_marker(ws.cell(row_idx, c).value) for c in duplicate_marker_cols)
        }
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in sheet_net_cols:
                value = ws.cell(row_idx, col_idx).value
                if nonempty(value):
                    nets.append((row_idx, col_idx, str(value).strip()))
            for col_idx in all_pin_cols:
                value = ws.cell(row_idx, col_idx).value
                if nonempty(value):
                    pins.append((row_idx, col_idx, str(value).strip()))

        net_counts = Counter(value for _, _, value in nets)
        for row_idx, col_idx, value in nets:
            if net_counts[value] > 1 and row_idx not in duplicate_rows:
                checks.append([ws.title, "duplicate_net_review", row_idx, get_column_letter(col_idx), value])

        pin_counts = Counter(value for _, _, value in pins)
        for row_idx, col_idx, value in pins:
            if pin_counts[value] > 1:
                checks.append([ws.title, "duplicate_pin_review", row_idx, get_column_letter(col_idx), value])

        if sheet_net_cols and all_pin_cols:
            if not source_cols:
                checks.append([ws.title, "missing_source_column_review", 1, "", ""])
            for row_idx in range(2, ws.max_row + 1):
                any_net = any(nonempty(ws.cell(row_idx, c).value) for c in sheet_net_cols)
                any_pin = any(nonempty(ws.cell(row_idx, c).value) for c in all_pin_cols)
                if any_net and not any_pin:
                    checks.append([ws.title, "net_without_pin_review", row_idx, "", ""])
                if any_pin and not any_net:
                    checks.append([ws.title, "pin_without_net_review", row_idx, "", ""])
                if (any_net or any_pin) and source_cols:
                    source_values = [
                        str(ws.cell(row_idx, c).value).strip()
                        for c in source_cols
                        if nonempty(ws.cell(row_idx, c).value)
                    ]
                    if not source_values:
                        checks.append([ws.title, "unsourced_row_review", row_idx, "", ""])
                    for value in source_values:
                        if value.lower() in SOURCE_REVIEW_MARKERS:
                            checks.append([ws.title, "source_status_review", row_idx, "", value])

        for row_idx in range(2, ws.max_row + 1):
            any_net = any(nonempty(ws.cell(row_idx, c).value) for c in sheet_net_cols)
            any_fpga_pin = any(nonempty(ws.cell(row_idx, c).value) for c in fpga_pin_cols)
            any_peripheral_pin = any(nonempty(ws.cell(row_idx, c).value) for c in peripheral_pin_cols)
            if any_net and fpga_pin_cols and not any_fpga_pin:
                checks.append([ws.title, "net_without_fpga_pin_review", row_idx, "", ""])
            if any_net and peripheral_pin_cols and not any_peripheral_pin:
                checks.append([ws.title, "net_without_peripheral_pin_review", row_idx, "", ""])
            if any_fpga_pin and peripheral_pin_cols and not any_peripheral_pin:
                checks.append([ws.title, "fpga_pin_without_peripheral_pin_review", row_idx, "", ""])
            if any_peripheral_pin and fpga_pin_cols and not any_fpga_pin:
                checks.append([ws.title, "peripheral_pin_without_fpga_pin_review", row_idx, "", ""])
    return checks


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--net-columns", help="Comma-separated net columns, e.g. B,D,F,H. Defaults to headers containing 'net'.")
    parser.add_argument("--pin-columns", help="Comma-separated pin/ball columns, e.g. A,C,E,G. Defaults to headers containing 'pin' or 'ball'.")
    parser.add_argument(
        "--skip-sheets",
        default=",".join(sorted(DEFAULT_SKIP_SHEETS)),
        help="Comma-separated sheet names to skip during mechanical checks.",
    )
    args = parser.parse_args()

    wb = load_workbook(args.input)
    net_cols = parse_cols(args.net_columns)
    pin_cols = parse_cols(args.pin_columns)
    skip_sheets = {name.strip() for name in args.skip_sheets.split(",") if name.strip()}

    for ws in wb.worksheets:
        style_sheet(ws)

    if "Mechanical_Checks" in wb.sheetnames:
        del wb["Mechanical_Checks"]
    ws_check = wb.create_sheet("Mechanical_Checks")
    ws_check.append(["Sheet", "Check", "Row", "Column", "Value"])
    checks = collect_checks(wb, net_cols, pin_cols, skip_sheets)
    if checks:
        for row in checks:
            ws_check.append(row)
    else:
        ws_check.append(["all", "no_mechanical_findings", "", "", ""])
    ws_check.append([])
    ws_check.append(["Generated", datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC"), "", "", ""])
    ws_check["A1"].fill = HEADER_FILL
    for cell in ws_check[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws_check["A3"].fill = NOTE_FILL
    style_sheet(ws_check)

    wb.save(args.output)
    print(f"Wrote {args.output}")
    print(f"Mechanical checks: {len(checks)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
