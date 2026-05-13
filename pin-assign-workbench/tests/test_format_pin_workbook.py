"""Tests for pin-assign-workbench/scripts/format_pin_workbook.py.

Run from repo root:

    python pin-assign-workbench/tests/test_format_pin_workbook.py

Or with pytest:

    python -m pytest pin-assign-workbench/tests/ -v
"""

from __future__ import annotations

import io
import sys
import tempfile
from contextlib import redirect_stdout
from pathlib import Path

from openpyxl import Workbook, load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
SCRIPT_DIR = REPO_ROOT / "pin-assign-workbench" / "scripts"
sys.path.insert(0, str(SCRIPT_DIR))

import format_pin_workbook  # noqa: E402


def _workbook_with_rows(headers: list[str], rows: list[list[object]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pin_Net_Output"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    return wb


def _check_names(checks: list[list[object]]) -> list[str]:
    return [str(row[1]) for row in checks]


def _collect(headers: list[str], rows: list[list[object]]) -> list[list[object]]:
    wb = _workbook_with_rows(headers, rows)
    return format_pin_workbook.collect_checks(
        wb,
        net_cols=set(),
        pin_cols=set(),
        skip_sheets=format_pin_workbook.DEFAULT_SKIP_SHEETS,
    )


def test_duplicate_net_is_reported():
    checks = _collect(
        ["Net", "FPGA Pin", "Source ID"],
        [
            ["DQ0", "A1", "S1"],
            ["DQ0", "A2", "S2"],
        ],
    )
    assert "duplicate_net_review" in _check_names(checks)


def test_duplicate_pin_is_reported():
    checks = _collect(
        ["Net", "FPGA Pin", "Source ID"],
        [
            ["DQ0", "A1", "S1"],
            ["DQ1", "A1", "S2"],
        ],
    )
    assert "duplicate_pin_review" in _check_names(checks)


def test_intentional_duplicate_net_marker_suppresses_duplicate_net_review():
    checks = _collect(
        ["Net", "FPGA Pin", "Source ID", "Intentional Duplicate"],
        [
            ["CLK", "A1", "S1", "Y"],
            ["CLK", "A2", "S2", "Y"],
        ],
    )
    assert "duplicate_net_review" not in _check_names(checks)


def test_signal_name_header_is_not_inferred_as_pin_column():
    headers = ["Signal Name", "Net", "Source ID"]
    assert format_pin_workbook.infer_generic_pin_columns(headers) == set()
    assert format_pin_workbook.infer_pin_columns(
        headers,
        format_pin_workbook.FPGA_PIN_SUBSTRINGS,
    ) == set()


def test_missing_source_column_is_reported():
    checks = _collect(
        ["Net", "FPGA Pin"],
        [["DQ0", "A1"]],
    )
    assert "missing_source_column_review" in _check_names(checks)


def test_unsourced_row_is_reported():
    checks = _collect(
        ["Net", "FPGA Pin", "Source ID"],
        [["DQ0", "A1", ""]],
    )
    assert "unsourced_row_review" in _check_names(checks)


def test_source_review_markers_are_reported():
    checks = _collect(
        ["Net", "FPGA Pin", "Source ID"],
        [
            ["DQ0", "A1", "TBD-source"],
            ["DQ1", "A2", "conflict"],
        ],
    )
    names = _check_names(checks)
    assert names.count("source_status_review") == 2


def test_clean_rows_have_no_collected_findings():
    checks = _collect(
        ["Net", "FPGA Pin", "Source ID"],
        [
            ["DQ0", "A1", "S1"],
            ["DQ1", "A2", "S2"],
        ],
    )
    assert checks == []


def test_main_writes_no_findings_and_is_idempotent():
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        source = tmp_path / "source.xlsx"
        first = tmp_path / "first.xlsx"
        second = tmp_path / "second.xlsx"

        wb = _workbook_with_rows(
            ["Net", "FPGA Pin", "Source ID"],
            [
                ["DQ0", "A1", "S1"],
                ["DQ1", "A2", "S2"],
            ],
        )
        wb.save(source)

        old_argv = sys.argv[:]
        try:
            sys.argv = ["format_pin_workbook.py", str(source), str(first)]
            with redirect_stdout(io.StringIO()):
                assert format_pin_workbook.main() == 0
            sys.argv = ["format_pin_workbook.py", str(first), str(second)]
            with redirect_stdout(io.StringIO()):
                assert format_pin_workbook.main() == 0
        finally:
            sys.argv = old_argv

        out = load_workbook(second)
        ws = out["Mechanical_Checks"]
        check_values = [
            ws.cell(row_idx, 2).value
            for row_idx in range(2, ws.max_row + 1)
            if ws.cell(row_idx, 2).value
        ]
        assert check_values.count("no_mechanical_findings") == 1


def _main() -> int:
    g = globals()
    tests = [(n, f) for n, f in g.items()
             if n.startswith("test_") and callable(f)]
    fails = 0
    for name, fn in tests:
        try:
            fn()
            print(f"PASS  {name}")
        except AssertionError as e:
            fails += 1
            print(f"FAIL  {name}: {e}")
        except Exception as e:
            fails += 1
            print(f"ERROR {name}: {type(e).__name__}: {e}")
    print(f"\n{len(tests) - fails}/{len(tests)} passed")
    return 1 if fails else 0


if __name__ == "__main__":
    sys.exit(_main())
